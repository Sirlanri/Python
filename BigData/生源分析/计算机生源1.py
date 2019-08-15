from openpyxl import load_workbook
import gb2260
from collections import Counter

wb=load_workbook(filename='信息.xlsx',read_only=True)
ws=wb['Sheet1']

mySheet=[]
after=[]
citys=[]

print(wb.sheetnames)#读取所有的sheet名称，保存为列表

#获得所有的身份证号
for row in ws.rows:
    mySheet.append(row[3].value)

#提取身份证号前四位（城市）
for one in mySheet:
    after.append(one[0:4])

for num in after:
    try:
        division=gb2260.get(num+'00')
        citys.append(division.name)
    except:
        citys.append('滨州市')#就你滨州事儿多
    finally:
        #将城市人数保存到字典place里
        place=Counter(citys)
print(place)


#开始尝试可视化
from pyecharts import options as opts
from pyecharts.charts import Bar
from pyecharts.charts import Pie

#试图排序的函数
def sort_by_value(d): 
    items=d.items() 
    backitems=[[v[1],v[0]] for v in items] 
    backitems.sort() 
    return [ backitems[i][1] for i in range(0,len(backitems))] 

def bar_base() ->Bar:
    c = (
        Bar()
        .add_xaxis(list(place.keys()))
        .add_yaxis("人数", list(place.values()))
        
        .set_global_opts(title_opts=opts.TitleOpts(title="18级新生来源", subtitle="by Python"))
    )
    c.render(path='直方图.html')


def pie_base() -> Pie:
    c = (
        Pie()
        .add("", [list(z) for z in zip(place.keys(), place.values())],
        radius=["40%", "75%"]
        )
        .set_global_opts(title_opts=opts.TitleOpts(title="18级新生来源",subtitle="by Python"))
        .set_series_opts(label_opts=opts.LabelOpts(formatter="{b}: {c}"))
    )
    c.render(path='饼图.html')

