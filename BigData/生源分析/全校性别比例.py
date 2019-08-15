from openpyxl import load_workbook
import gb2260
from collections import Counter
from pyecharts import options as opts
from pyecharts.charts import Bar
from pyecharts.charts import Pie

#设置全局变量
allNum=[]#完整的身份证号
placeNum=[]#区域代码，前6位
cities=Counter()#城市出现次数
boyNum=girlNum=0#男女数量
percent=0.0
allPercent={}#所有学院对应的男女比例
seriseNames=[]#所有学院名字
sexList=[]#所有学院男女比的值
seriseLen=[]#记录各个学院的数据长度
i=0

#从表格中获取身份证号
def getNum(ExcelName):
    global allNum,seriseNames
    wb=load_workbook(filename=ExcelName,read_only=True)
    allSheet=wb.sheetnames#获得所有的sheet
    for sheet in allSheet:
        seriesSheet=wb[sheet]#读取各个学院的sheet
        seriseNames.append(sheet)#获得所有学院名字
        for row in seriesSheet.rows:
            allNum.append(row[1].value)#把身份证号码添加到列表
        seriseLen.append(len(allNum))


#判断性别
def sex(num):
    global girlNum,boyNum
    '''接收的是完整的身份证号'''
    try:
        if num!=None and len(num)==18:
            if int(num[16:17])%2==0:
                girlNum+=1
            else:
                boyNum+=1
    except:
        print("性别判断出错")

#统计生源地区（城市）数量
def place(num):
    '''接收的是完整的身份证号'''
    global cities
    city=[]
    try:
        division=gb2260.get(num[0:6])
        city.append(division.name)
    except:
        print('判断城市错误'+num)
    finally:
        cities=Counter(city)
        print('\n城市统计完成')

#汇总单个学院男女比例
#全校比例的计算出了问题

def sexPercent():
    global girlNum,boyNum,percent,allPercent
    global i
    sexList.append(boyNum/girlNum)
    for seriseName in seriseNames:
        if i==0:
            allPercent[seriseName]=sexList[0:seriseLen[0]]
            
        else:
            try:
                allPercent[seriseName]=sexList[i]
            except:
                print('比例出错')
        
        print(seriseName+str(sexList))
        i+=1
            

def do_sex():
    global allPercent
    getNum('各学院.xlsx')
    for num in allNum:
        sex(num)
    sexPercent()
    print(allPercent)

#可视化
def vision():
    c=(
        Pie()
        .add('',boyNum,girlNum)
        .set_global_opts(title_opts=opts.TitleOpts(title="18级男女比例",subtitle="by Python"))
        .set_series_opts(label_opts=opts.LabelOpts(formatter="{b}: {c}"))
    )
    c.render(path='男女比例.html')

do_sex()